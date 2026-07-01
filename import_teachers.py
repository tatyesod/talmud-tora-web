#!/usr/bin/env python3
"""ייבוא רשימת עובדים מ-Excel — מוחק הכל ומייבא מחדש"""
import openpyxl, sqlite3, os, warnings, datetime
warnings.filterwarnings('ignore')

DATA_DIR = os.environ.get("RENDER_PERSISTENT_DIR", os.path.join(os.path.dirname(__file__), "server", "data"))
DB_PATH = os.path.join(DATA_DIR, "talmud-tora.db")
XLS_PATH = os.path.join(os.path.dirname(__file__), "teacher_data.xlsx")

if not os.path.exists(XLS_PATH):
    print(f"שגיאה: לא נמצא קובץ {XLS_PATH}")
    exit(1)

wb = openpyxl.load_workbook(XLS_PATH)
ws = wb.active

conn = sqlite3.connect(DB_PATH)
cur = conn.cursor()

# וידוא עמודות חדשות קיימות
extra_cols = [
    "ALTER TABLE teachers ADD COLUMN id_number_spouse TEXT",
    "ALTER TABLE teachers ADD COLUMN spouse_last_name TEXT",
    "ALTER TABLE teachers ADD COLUMN spouse_first_name TEXT",
    "ALTER TABLE teachers ADD COLUMN health_fund TEXT",
    "ALTER TABLE teachers ADD COLUMN email TEXT",
    "ALTER TABLE teachers ADD COLUMN children_count_total INTEGER",
    "ALTER TABLE teachers ADD COLUMN gender TEXT",
    "ALTER TABLE teachers ADD COLUMN branch TEXT",
]
for sql in extra_cols:
    try: cur.execute(sql)
    except: pass

# מחיקת כל הרשומות הקיימות
cur.execute("DELETE FROM teacher_classes")
cur.execute("DELETE FROM teacher_file")
cur.execute("DELETE FROM teacher_attendance")
cur.execute("DELETE FROM teachers")
print("נמחקו כל הרשומות הקיימות")

# קריאת headers
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
print("עמודות:", headers)

def parse_phone(v):
    if not v: return None
    s = str(v).strip().replace('-','').replace(' ','')
    return s if s else None

def parse_date(v):
    if not v: return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        # convert to access-style serial (days since 1899-12-30)
        epoch = datetime.date(1899, 12, 30)
        d = v.date() if hasattr(v, 'date') else v
        return (d - epoch).days
    return None

count = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    if not any(row): continue
    
    branch    = row[0]
    id_num    = str(row[1]).strip() if row[1] else None
    last_name = row[2]
    first_name= row[3]
    birth_date= parse_date(row[4])
    street    = row[5]
    house_num = str(row[6]).strip() if row[6] else None
    city      = row[7]
    phone     = parse_phone(row[8])
    mobile    = parse_phone(row[9])
    gender    = row[10]
    health_fund = row[11]
    email     = row[12]
    children  = row[13]
    id_spouse = str(row[14]).strip() if row[14] else None
    sp_last   = row[15]
    sp_first  = row[16]

    cur.execute("""
        INSERT INTO teachers (
            last_name, first_name, id_number, birth_date_civil,
            street, house_number, city, home_phone, mobile,
            gender, health_fund, email, children_count_total,
            id_number_spouse, spouse_last_name, spouse_first_name,
            branch, status
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        last_name, first_name, id_num, birth_date,
        street, house_num, city, phone, mobile,
        gender, health_fund, email, children,
        id_spouse, sp_last, sp_first,
        branch, 'פעיל'
    ))
    count += 1

conn.commit()
conn.close()
print(f"יובאו {count} עובדים")
